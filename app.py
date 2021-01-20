from flask import Flask, jsonify, request, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import sys
import uuid
import cv2
import numpy as np
import pandas as pd
from pdf2image import convert_from_path
from sklearn.cluster import KMeans
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as Image_xl
from PIL import Image
import itertools
from functools import reduce
import shutil
from sklearn.externals import joblib
from numba import cuda
from datetime import datetime
import tempfile
import docx
from xml.etree import ElementTree
import xml.etree.cElementTree as ET
from io import StringIO
import sqlite3
from win32com import client
import win32api
import pythoncom

app = Flask(__name__)
APP_ROOT = os.path.dirname(os.path.abspath(__file__))
CORS(app)

dbPath = sys.path[0]+'/database.db'
cx = sqlite3.connect(dbPath, check_same_thread=False)

# 絕對路徑自行設定
pdfsPath = 'D:\\project\\PhotoCompare\\backend\\pdfs'
wordsPath = 'D:\\project\\PhotoCompare\\backend\\words'
excelsPath = 'D:\\project\\PhotoCompare\\backend\\excels'
rawImgsPath = 'D:\\project\\PhotoCompare\\backend\\raw_imgs'
resizeImgsPath = 'D:\\project\\PhotoCompare\\frontend\\public\\static\\resize_imgs'
# resizeImgsPath = 'D:\\project\\PhotoCompare\\frontend\\dist\\static\\resize_imgs'
modelsPath = 'D:\\project\\PhotoCompare\\backend\\models'
resultsPath = 'D:\\project\\PhotoCompare\\backend\\results'
excelResultsPath = 'D:\\project\\PhotoCompare\\backend\\excel_results'


@app.route("/login", methods=['POST'])
def login():
    response = {"success": False}
    if request.method == 'POST':
        dcit_request = request.get_json()
        username = dcit_request['username']
        password = dcit_request['password']
        cu = cx.cursor()
        try:
            cu.execute(
                "select password from 'User' where username=\'" + username + "\'")
            pwd = cu.fetchone()[0]
        except:
            pwd = None
        cu.close()
        if not pwd:
            return "-1"
        elif password != pwd:
            return "0"
        elif username == "admin" and password == pwd:
            return "admin"
        else:
            return "1"
    else:
        return jsonify(response)

@app.route("/create", methods=["POST"])
def create():
    post_data = request.get_json()
    username = post_data.get('username')
    path = os.path.join(APP_ROOT, username)
    if not os.path.isdir(path):
        os.mkdir(path)
    response_object = {'status': 'success'}
    return jsonify(response_object)

@app.route("/destroy", methods=["POST"])
def destroy():
    post_data = request.get_json()
    username = post_data.get('username')
    path = os.path.join(APP_ROOT, username)
    if os.path.isdir(path):
        shutil.rmtree(path)
    response_object = {'status': 'success'}
    return jsonify(response_object)


@app.route("/users", methods=["GET", "POST"])
def users():
    response_object = {'status': 'success'}
    if request.method == 'GET':
        cu = cx.cursor()
        cu.execute("select username, password, remark from 'User'")
        data = cu.fetchall()
        cu.close()
        userList = []
        for d in data:
            userList.append(
                {"username": d[0], "password": d[1], "remark": d[2]})
        response_object['users'] = userList
    if request.method == 'POST':
        post_data = request.get_json()
        username = post_data.get('username')
        password = post_data.get('password')
        remark = ''
        remark = post_data.get('remark')
        cu = cx.cursor()
        cu.execute("insert into User values(null, '" + username +
                   "', '" + password + "', '" + remark + "')")
        cu.close()
        cx.commit()
        response_object['message'] = 'User added!'

    return jsonify(response_object)


@app.route('/user/<user_id>', methods=['PUT', 'DELETE'])
def user(user_id):
    response_object = {'status': 'success'}
    if request.method == 'PUT':
        post_data = request.get_json()
        username = post_data.get('username')
        password = post_data.get('password')
        remark = ''
        remark = post_data.get('remark')
        cu = cx.cursor()
        cu.execute("UPDATE User SET password = '" + password +
                   "', remark = '" + remark + "' WHERE username = '" + username + "'")
        cu.close()
        cx.commit()
        response_object['message'] = 'User updated!'
    if request.method == 'DELETE':
        cu = cx.cursor()
        cu.execute("DELETE FROM User WHERE username = '" + user_id + "'")
        cu.close()
        cx.commit()
        response_object['message'] = 'User removed!'
    return jsonify(response_object)


@app.route('/imgs', methods=["GET"])
def imgs():
    response_object = {'status': 'success'}
    if request.method == 'GET':
        cu = cx.cursor()
        cu.execute("select imageName, uploadUser from 'Image'")
        data = cu.fetchall()
        cu.close()
        imgList = []
        for d in data:
            imgList.append({"imageName": d[0], "uploadUser": d[1]})
        response_object['images'] = imgList
    return jsonify(response_object)


@app.route('/delimg/<img>', methods=['DELETE'])
def delimg(img):
    response_object = {'status': 'success'}
    if request.method == 'DELETE':
        cu = cx.cursor()
        cu.execute("DELETE FROM Image WHERE imageName = '" + img + "'")
        cu.close()
        cx.commit()
        response_object['message'] = 'Image removed!'
    return jsonify(response_object)


@app.route('/uploadrecord', methods=['GET'])
def uploadrecord():
    response_object = {'status': 'success'}
    if request.method == 'GET':
        cu = cx.cursor()
        cu.execute(
            "select time, fileName, uploadUser, result from 'UploadRecord' order by time DESC")
        data = cu.fetchall()
        cu.close()
        cx.commit()
        list = []
        for d in data:
            list.append({"time": d[0], "fileName": d[1],
                         "uploadUser": d[2], 'result': d[3]})
        response_object['uploadRecordList'] = list
    return jsonify(response_object)


@app.route('/deletrecord', methods=['POST'])
def deleterecord():
    response_object = {'status': 'success'}
    if request.method == 'POST':
        post_data = request.get_json()
        fileName = post_data.get('fileName')
        result = post_data.get('result')
        os.remove(os.path.join(resultsPath, result))
        cu = cx.cursor()
        cu.execute("DELETE FROM UploadRecord WHERE result = '" + result + "'")
        cu.close()
        cx.commit()
    return jsonify(response_object)


@app.route('/upload', methods=['POST'])
def upload():
    response = {"success": False}
    try:
        if request.method == 'POST':
            post_data = request.get_json()
            formData = post_data['formData']
            username = post_data['username']
            print(username)
            files = formData.files
            print(len(files))
            # files = request.files
            # if len(files) != 0:
            #     autoGenFolderName = uuid.uuid1()
            #     folderPath = os.path.join(APP_ROOT, str(autoGenFolderName))
            #     os.mkdir(folderPath)
            # for f in files:
            #     file = request.files[f]
            #     filename = secure_filename(file.filename)
            #     file.save(os.path.join(folderPath, filename))
            response = {
                "success": True,
                "folderPath": folderPath,
            }
    except Exception as e:
        print(e)

    return jsonify(response)


@app.route('/delete', methods=['POST'])
def delete():
    dcit_request = request.get_json()
    path = os.path.join(dcit_request['folderPath'], dcit_request['file'])
    os.remove(path)
    fileList = os.listdir(dcit_request['folderPath'])
    if len(fileList) == 0:
        shutil.rmtree(dcit_request['folderPath'])
    response = {"success": True}
    return jsonify(response)


@app.route('/getlist', methods=['POST'])
def uploadList():
    print(request.get_json())
    dcit_request = request.get_json()
    folderPath = dcit_request['folderPath']
    if os.path.isdir(folderPath):
        fileList = os.listdir(folderPath)
    else:
        fileList = []
    response = {
        "fileList": fileList,
    }
    return jsonify(response)

@app.route('/download', methods=['POST'])
def download():
    dcit_request = request.get_json()
    resultFileName = dcit_request['resultFileName']
    f = open(os.path.join(resultsPath, resultFileName), 'rb')
    return f.read()

@app.route('/updatedb', methods=['POST'])
def updatedb():
    dcit_request = request.get_json()
    username = dcit_request['username']
    cu = cx.cursor()
    for imgData in dcit_request['imgsData']:
        cu.execute("SELECT * FROM Image WHERE imageName = '" +
                   imgData['imgName'] + "'")
        data = cu.fetchall()
        if len(data) == 0:
            cu.execute("insert into Image values(null, '" +
                       imgData['imgName'] + "', '" + imgData['pHash'] + "', '" + imgData['group'] + "', '" + username + "')")
        else:
            cu.execute("update Image SET pHash = '" + imgData['pHash'] + "', g = '" + imgData['group'] +
                       "', uploadUser = '" + username + "' WHERE imageName like '" + imgData['imgName'] + "'")
    cu.close()
    cx.commit()
    response = {"success": True}
    return jsonify(response)


@app.route('/getImg', methods=['POST'])
def getImg():
    dcit_request = request.get_json()
    folderPath = dcit_request['folderPath']
    files = os.listdir(folderPath)
    imgsPath = folderPath+'_imgs'
    if not os.path.isdir(imgsPath):
        os.mkdir(imgsPath)
    filesName = []
    for file in files:
        if str(file).split('.')[-1].lower() == 'docx':
            getWordImgs(folderPath, file, imgsPath)
            shutil.copy(os.path.join(folderPath, file),
                        os.path.join(wordsPath, file))
        elif str(file).split('.')[-1].lower() == 'xlsx':
            getExcelImgs(folderPath, file, imgsPath)
            shutil.copy(os.path.join(folderPath, file),
                        os.path.join(excelsPath, file))
        elif str(file).split('.')[-1].lower() == 'pdf':
            getPDFImgs(folderPath, file, imgsPath)
            shutil.copy(os.path.join(folderPath, file),
                        os.path.join(pdfsPath, file))
        fileName = str(file).split('_')[0]
        if fileName not in filesName:
            filesName.append(fileName)

    imgsNum = len(os.listdir(imgsPath))

    response = {
        "success": True,
        "imgsPath": imgsPath,
        "imgsNum": imgsNum
    }
    return jsonify(response)


@app.route('/compare', methods=['POST'])
def compare():

    dcit_request = request.get_json()
    fileList = dcit_request['fileList']
    folderPath = dcit_request['folderPath']
    imgsPath = dcit_request['imgsPath']

    # 將相片轉為特徵
    imgsName = os.listdir(imgsPath)
    imgsPHash = {}
    imgsFingerprint = {}
    for imgName in imgsName:
        img = Image.open(os.path.join(imgsPath, imgName))
        pHash = phash(img)
        imgsPHash[imgName] = pHash
        fingerprint = [bool(int(d)) for d in str(bin(pHash))[2:].zfill(1024)]
        imgsFingerprint[imgName] = fingerprint
        shutil.move(os.path.join(imgsPath, imgName),
                    os.path.join(rawImgsPath, imgName))

    shutil.rmtree(folderPath)
    shutil.rmtree(imgsPath)

    # 上傳相片倆倆比對
    result1 = []
    duplicateImgs = []
    for imgName1, imgName2 in itertools.combinations(imgsName, 2):
        pHash1 = imgsPHash[imgName1]
        pHash2 = imgsPHash[imgName2]
        distance = bin(pHash1 ^ pHash2).count('1')
        similary = 1 - distance / 1024
        if similary > 0.9:
            if imgName1 not in duplicateImgs:
                result1.append(
                    {'imgName1': imgName1, 'imgName2': imgName2, 'similary': similary})
                duplicateImgs.append(imgName2)

    # 將上傳相片分群
    kmeans = joblib.load(os.path.join(modelsPath, 'KMeans_model.m'))
    pred = kmeans.predict(list(imgsFingerprint.values()))
    imgsGroup = dict(zip(list(imgsFingerprint.keys()), pred))
    imgsGroupDF = pd.DataFrame(
        data=list(imgsFingerprint.keys()), columns=['imageName'])
    imgsGroupDF['group'] = pred

    # 上傳相片與資料庫相片比對
    result2 = []
    cu = cx.cursor()
    for group in set(imgsGroupDF.group):
        imgsNm = imgsGroupDF.query("group == {}".format(group)).imageName
        cu.execute(
            "select imageName, pHash from Image WHERE g like '" + str(group) + "'")
        data = cu.fetchall()
        dbImages = []
        for d in data:
            dbImages.append([d[0], d[1]])
        for imgName1 in imgsNm:
            pHash1 = imgsPHash[imgName1]
            for dbImage in dbImages:
                imgName2 = dbImage[0]
                pHash2 = int(dbImage[1])
                distance = bin(pHash1 ^ pHash2).count('1')
                similary = 1 - distance / 1024
                if similary > 0.9:
                    if imgName1 not in duplicateImgs:
                        result2.append(
                            {'imgName1': imgName1, 'imgName2': imgName2, 'similary': similary})
                        if imgName1 not in duplicateImgs:
                            duplicateImgs.append(imgName1)
    cu.close()

    wb = Workbook()
    wb.create_sheet("工作表1", 0)
    sht = wb['工作表1']
    sht.merge_cells('A1:E1')
    sht['A1'] = '工程案件相片重複性辨識'
    sht['A1'].font = Font(size=16, b=True, underline='single')
    sht['A1'].alignment = Alignment(horizontal='center', vertical='center')
    fileName = fileList[0].split('.')[0]
    sht['A2'] = '工程檔案名稱：' + fileName
    sht['E2'] = '日期：' + datetime.now().strftime("%Y/%m/%d")
    sht['A2'].font = Font(size=14, b=True)
    sht['E2'].font = Font(size=14, b=True)
    sht.column_dimensions["A"].width = 40
    sht.column_dimensions["B"].width = 25
    sht.column_dimensions["C"].width = 40
    sht.column_dimensions["D"].width = 25
    sht.column_dimensions["E"].width = 25

    sht.merge_cells('A3:E3')
    sht['A3'].font = Font(size=14, b=True)
    if len(duplicateImgs) == 0:
        message = '[系統比對結果]相片無重複，是否將無重複的 ' + \
            str(len(imgsName)) + ' 張相片寫入資料庫？(寫入後才能出表)'
        sht['A3'] = '[系統比對結果]無重複相片(系統比對相片數量：' + str(len(imgsName)) + \
            '張，寫入資料庫相片數量：' + str(len(imgsName)) + '張)'
    elif len(imgsName) == len(duplicateImgs):
        message = '[系統比對結果]全部相片重複'
        sht['A3'] = '[系統比對結果]全部相片重複(系統比對相片數量：' + str(len(imgsName)) + '張)'
    else:
        message = '[系統比對結果]部分相片重複，是否將無重複的 ' + \
            str(len(imgsName) - len(duplicateImgs)) + ' 張相片寫入資料庫？'
        sht['A3'] = '[系統比對結果]部分相片重複，重複相片數量：' + \
            str(len(duplicateImgs)) + '張(系統比對相片數量：' + str(len(imgsName)) + '張)'


    line = Side(style='thin', color='000000')
    border = Border(top=line, bottom=line, left=line, right=line)

    i = 4
    if len(result1) != 0:
        sht.merge_cells('A'+str(i)+':B'+str(i))
        sht['A'+str(i)] = '上傳的相片'
        sht['A'+str(i)].font = Font(b=True)
        sht['A'+str(i)].border = border
        sht['A'+str(i)].alignment = Alignment(horizontal='center',
                                              vertical='center')

        sht.merge_cells('C'+str(i)+':D'+str(i))
        sht['C'+str(i)] = '上傳的相片'
        sht['C'+str(i)].font = Font(b=True)
        sht['C'+str(i)].border = border
        sht['C'+str(i)].alignment = Alignment(horizontal='center',
                                              vertical='center')
        sht['E'+str(i)] = '說明/備註'
        sht['E'+str(i)].font = Font(b=True)
        sht['E'+str(i)].border = border
        sht['E'+str(i)].alignment = Alignment(horizontal='center',
                                              vertical='center')
        i = i+1
        for item in result1:
            sht['A'+str(i)].border = border
            sht['B'+str(i)].border = border
            sht['C'+str(i)].border = border
            sht['D'+str(i)].border = border
            sht['E'+str(i)].border = border
            sht.row_dimensions[i].height = 80
            with Image.open(os.path.join(rawImgsPath, item['imgName1'])) as img:
                img = img.resize((100, 100), Image.ANTIALIAS)
                img.save(os.path.join(resizeImgsPath, item['imgName1']))
            with Image.open(os.path.join(rawImgsPath, item['imgName2'])) as img:
                img = img.resize((100, 100), Image.ANTIALIAS)
                img.save(os.path.join(resizeImgsPath, item['imgName2']))
            sht['A'+str(i)] = item['imgName1']
            sht.add_image(Image_xl(os.path.join(
                resizeImgsPath, item['imgName1'])), "B"+str(i))
            sht['C'+str(i)] = item['imgName2']
            sht.add_image(Image_xl(os.path.join(
                resizeImgsPath, item['imgName2'])), "D"+str(i))
            i = i + 1

    if len(result2) != 0:
        sht.merge_cells('A'+str(i)+':B'+str(i))
        sht['A'+str(i)] = '上傳的相片'
        sht['A'+str(i)].font = Font(b=True)
        sht['A'+str(i)].border = border
        sht['A'+str(i)].alignment = Alignment(horizontal='center',
                                              vertical='center')
        sht.merge_cells('C'+str(i)+':D'+str(i))
        sht['C'+str(i)] = '資料庫的相片'
        sht['C'+str(i)].font = Font(b=True)
        sht['C'+str(i)].border = border
        sht['C'+str(i)].alignment = Alignment(horizontal='center',
                                              vertical='center')
        sht['E'+str(i)] = '說明/備註'
        sht['E'+str(i)].font = Font(b=True)
        sht['E'+str(i)].border = border
        sht['E'+str(i)].alignment = Alignment(horizontal='center',
                                              vertical='center')
        i = i+1
        for item in result2:
            sht['A'+str(i)].border = border
            sht['B'+str(i)].border = border
            sht['C'+str(i)].border = border
            sht['D'+str(i)].border = border
            sht['E'+str(i)].border = border
            sht.row_dimensions[i].height = 80
            with Image.open(os.path.join(rawImgsPath, item['imgName1'])) as img:
                img = img.resize((100, 100), Image.ANTIALIAS)
                img.save(os.path.join(resizeImgsPath, item['imgName1']))
            with Image.open(os.path.join(rawImgsPath, item['imgName2'])) as img:
                img = img.resize((100, 100), Image.ANTIALIAS)
                img.save(os.path.join(resizeImgsPath, item['imgName2']))
            sht['A'+str(i)] = item['imgName1']
            sht.add_image(Image_xl(os.path.join(
                resizeImgsPath, item['imgName1'])), "B"+str(i))
            sht['C'+str(i)] = item['imgName2']
            sht.add_image(Image_xl(os.path.join(
                resizeImgsPath, item['imgName2'])), "D"+str(i))
            i = i + 1

    sht.page_setup.orientation = sht.ORIENTATION_LANDSCAPE
    sht.page_setup.fitToHeight = False
    sht.sheet_properties.pageSetUpPr.fitToPage = True

    resultFileName = datetime.today().strftime(
        "%Y%m%d%H%M") + '_' + fileName + '_比對結果.xlsx'
    wb.save(os.path.join(excelResultsPath, resultFileName))
    wb.close()

    pythoncom.CoInitialize()
    app = client.Dispatch("Excel.Application")
    pythoncom.CoInitialize()
    wb = app.Workbooks.Open(os.path.join(excelResultsPath,  resultFileName))
    ws = wb.Worksheets[0]
    ws.Visible = 1
    resultFileName = datetime.today().strftime(
        "%Y%m%d%H%M") + '_' + fileName + '_比對結果.pdf'
    ws.ExportAsFixedFormat( 0, os.path.join(resultsPath, resultFileName))
    wb.Close()
    app.Quit()

    s1 = set(imgsName)
    s2 = set(duplicateImgs)
    nonDuplicateImgs = list(s1.symmetric_difference(s2))

    nonDuplicateImgsData = []
    for nonDuplicateImg in nonDuplicateImgs:
        pHash = str(imgsPHash[nonDuplicateImg])
        group = str(imgsGroup[nonDuplicateImg])
        nonDuplicateImgsData.append(
            {'imgName': nonDuplicateImg, 'pHash': pHash, 'group': group})

    cu = cx.cursor()
    for file in fileList:
        cu.execute("insert into UploadRecord values(null, '" + datetime.now().strftime(
            "%Y/%m/%d %H:%M") + "', '" + file + "', '" + dcit_request['username'] + "', '" + resultFileName + "')")
    cu.close()
    cx.commit()

    response = {
        "success": True,
        "result1": result1,
        "result2": result2,
        "resultFileName": resultFileName,
        "duplicateImgsNum": len(duplicateImgs),
        "message": message,
        "nonDuplicateImgsData": nonDuplicateImgsData,
    }
    return jsonify(response)


def phash(img):
    img = img.resize((32, 32), Image.ANTIALIAS).convert('L')
    avg = reduce(lambda x, y: x + y, img.getdata()) / 1024.
    hash_value = reduce(lambda x, y: x | (y[1] << y[0]), enumerate(
        map(lambda i: 0 if i < avg else 1, img.getdata())), 0)
    return hash_value


def getWordImgs(folderPath, file, imgsPath):
    doc = docx.Document(os.path.join(folderPath, file))
    tables = doc.tables
    page = 0
    for table in tables:
        xml_str = table._element.xml
        root = ET.fromstring(xml_str)
        namespaces = dict([node for _, node in ElementTree.iterparse(
            StringIO(xml_str), events=['start-ns'])])
        i = 1
        try:
            for blip_elem in root.findall('.//a:blip', namespaces):
                embed_attr = blip_elem.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                img = doc.part.related_parts[embed_attr]
                page_num = (str(page+1)).zfill(3)
                dst_filenm = '_'.join([file, 'Page'+page_num, str(i)])+'.jpg'
                with open(os.path.join(imgsPath, dst_filenm), 'wb') as f:
                    f.write(img.blob)
                i += 1
            for imagedata_elem in root.findall('.//v:imagedata', namespaces):
                id_attr = imagedata_elem.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                img = doc.part.related_parts[id_attr]
                page_num = (str(page+1)).zfill(3)
                dst_filenm = '_'.join([file, 'Page'+page_num, str(i)])+'.jpg'
                with open(os.path.join(imgsPath, dst_filenm), 'wb') as f:
                    f.write(img.blob)
                i += 1
            page += 1
        except:
            pass


def getExcelImgs(folderPath, file, imgsPath):
    wb = load_workbook(os.path.join(folderPath, file))
    sheets = wb.get_sheet_names()
    for sheet in sheets:
        ws = wb[sheet]
        for i, image in enumerate(ws._images):
            img = Image.open(image.ref)
            img = img.convert("RGB")
            imgFileName = '_'.join([file, sheet, str(i+1)])+'.jpg'
            img.save(os.path.join(imgsPath, imgFileName))

def getPDFImgs(folderPath, file, imgsPath):
    pageImgs = convert_from_path(os.path.join(
        folderPath, file), output_folder=folderPath, dpi=600, fmt='jpg')

    for pageNumber, pageImg in enumerate(pageImgs):
        img = np.array(pageImg)
        pageImg.close()
        rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
        # gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # sat
        # 取出飽和度
        saturation = hsv[:, :, 1]
        _, threshold = cv2.threshold(
            saturation, 1, 255.0, cv2.THRESH_BINARY)
        # 2值化圖去除雜訊
        kernel_radius = int(threshold.shape[1]/300)
        kernel = np.ones((kernel_radius, kernel_radius), np.uint8)
        threshold = cv2.morphologyEx(threshold, cv2.MORPH_OPEN, kernel)
        # 產生等高線
        contours, hierarchy = cv2.findContours(
            threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        sat_conts = [
            contour for contour in contours if cv2.contourArea(contour) > 500000]

        conts = sat_conts

        sortY_conts = sorted([cont for cont in conts],
                                key=lambda x: x[0][0][1], reverse=False)
        up_conts = sortY_conts[:3]
        up_conts = sorted([cont for cont in up_conts],
                            key=lambda x: x[0][0][0], reverse=False)
        down_conts = sortY_conts[3:]
        down_conts = sorted([cont for cont in down_conts],
                            key=lambda x: x[0][0][0], reverse=False)
        merge_conts = up_conts+down_conts

        for i, c in enumerate(merge_conts):
            # 嘗試在各種角度，以最小的方框包住面積最大的等高線區域，以紅色線條標示
            rect = cv2.minAreaRect(c)
            box = cv2.boxPoints(rect)
            box = np.int0(box)
            angle = rect[2]
            if angle < -45:
                angle = 90 + angle
            # 以影像中心為旋轉軸心
            (h, w) = img.shape[:2]
            center = (w // 2, h // 2)
            # 計算旋轉矩陣
            M = cv2.getRotationMatrix2D(center, angle, 1.0)
            # 旋轉圖片
            rotated = cv2.warpAffine(
                rgb, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_CONSTANT)
            # 旋轉紅色方框座標
            pts = np.int0(cv2.transform(np.array([box]), M))[0]
            #  計算旋轉後的紅色方框範圍
            y_min = min(pts[0][0], pts[1][0], pts[2][0], pts[3][0])
            y_max = max(pts[0][0], pts[1][0], pts[2][0], pts[3][0])
            x_min = min(pts[0][1], pts[1][1], pts[2][1], pts[3][1])
            x_max = max(pts[0][1], pts[1][1], pts[2][1], pts[3][1])
            # 裁切影像
            img_crop = rotated[x_min:x_max, y_min:y_max]
            page_num = (str(pageNumber+1)).zfill(3)
            dst_filenm = '_'.join([file, 'Page'+page_num, str(i+1)])+'.jpg'
            cv2.imwrite(os.path.join(imgsPath, dst_filenm), img_crop)


if __name__ == "__main__":
    app.run(debug=False)
